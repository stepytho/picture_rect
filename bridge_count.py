import tkinter as tk
import cong_num
from PIL import Image, ImageGrab 
import screen_yes
import write
from openpyxl import Workbook, load_workbook
import pyperclip
import _thread
import time
import keyboard

#开关
def from_left():
    global change
    change = True
def from_right():
    global change
    change = False


def continue_recogninze():
    global continue_rec
    if var_check.get() == 0:
        continue_rec = False
        print("关闭连续识别")
    elif var_check.get() == 1:
        continue_rec = True
        print("打开连续识别")

#识别模块
#识别函数
def recognize():
    var_show.set("")
    nunbers = cong_num.main()
    if not continue_rec:
        text1.delete('1.0', tk.END)
    for i in nunbers:
        text1.insert("insert", i+"\n")
def recognize_angle():
    var_show.set("")
    nunbers = cong_num.main_angle()
    if not continue_rec:
        text2.delete('1.0', tk.END)
    for i in nunbers:
        text2.insert("insert", i+"\n")


#计算模块
text_content = []#用来储存每一行内容的列表
final_num = []#最终纯原始数字列表
final_num_1 = []#除2算法用表
final_num_angle = []
result_num = []#用于显示
result_num_1 = []#
result_num_angle = []#用于显示
result_num_write = []#用于写入
result_num_angle_write = []#用于写入
compare1 = []#用于比较
compare2 = []#用于比较

#一键清空
def empty():
    text1.delete('1.0', tk.END)
    text2.delete('1.0', tk.END)
    text3.delete('1.0', tk.END)
    text4.delete('1.0', tk.END)
    var_show.set("")
    var_show.set("--已全部清空--")
    print("--已全部清空--")

#快捷键函数
def rect(delay,key,name):

    while True:
        time.sleep(delay)
        if name == "empty":
            keyboard.wait("Ctrl+"+key)
            empty()
        elif name == "count":
            keyboard.wait("Ctrl+"+key)
            calculate()
            calculate_angle()
            var_show.set("")
            var_show.set("--计算成功!--")
            print("计算完成")       
        elif name == "left" or "right":
            keyboard.wait("Ctrl+"+key)
            cut(name)

        #elif name == "copy":
            #keyboard.wait("Ctrl+"+key)
            #print("--复制成功--")

        

def rec_cut(delay,key,name):
    
    while True:
        time.sleep(delay)
        keyboard.wait("Ctrl+"+key)
        print("截屏",name)
        cut(name)

#截屏函数
def cut(name):
    screen_yes.cut()
    if name == "left":
        recognize()
        var_show.set("")
        var_show.set("--横向截图，识别完成--")
    elif name == "right":
        recognize_angle()
        var_show.set("")
        var_show.set("--纵向截图，识别完成--")        
    

#获取每一行数值
def getData():
    get_resluts(text1,final_num)
def getData_angle():
    get_resluts(text2,final_num_angle)



#计算位置
def sum_calculate():
    sum = 0
    #print("final_num = ",final_num)
    result_num.clear()
    if change:
        final_num.reverse()
    for i in final_num:
        try:
            sum += int(i)
        except ValueError:
            sum += float(i)
        finally:
            result_num.insert(0,"-"+str(sum))
            result_num.append(str(sum))
        #print("sum =",sum)
    #print(result_num)

def sum_calculate_angle():

    result_num_angle.clear()
    if change:
        final_num_angle.reverse()
    for i in final_num_angle:
        result_num_angle.insert(0,"-"+i)
        result_num_angle.append("-"+i)

    #print("竖弯列表是",result_num_angle)        


def calculate():
    getData()
    sum_calculate()
    text3.delete('1.0', tk.END)
    for i in result_num:
        text3.insert("insert", i+"\n")
    
    #print("此时横向尺寸列表为", result_num)
def tran(list_):
    try:
        if int(list_[-1])%2 == 0:
            list_[-1] = str(int(int(list_[-1])/2)) 
        else:
            list_[-1] = str(int(list_[-1])/2)
    except ValueError:
        list_[-1] = str(float(list_[-1])/2)
    return list_

def divide():
    final_num_1.clear()
    old_list = get_resluts(text1,final_num_1)
    now_list = tran(old_list)
    text1.delete('1.0', tk.END)
    for i in now_list:
        text1.insert("insert", i+"\n")

def calculate_angle():
    getData_angle()
    sum_calculate_angle()
    text4.delete('1.0', tk.END)
    for i in result_num_angle:
        text4.insert("insert", i+"\n")

    #print("此时竖向尺寸列表为", result_num_angle)

#获取Text部件里内容函数
def get_resluts(text,list_):
    text_content = (text.get("0.0","end").replace(" ","")).split("\n")
    text_content.pop()#列表最后一个元素是空删除它
    list_.clear()
    for i in text_content:
        if i != "":
            list_.append(i)
    #print("获取到text框里的列表为",list_)
    return list_

#复制到粘贴板
def copy_datas(name,text):
    text_content = (text.get("0.0","end").replace(" ","")).split("\n")
    text_content.pop()
    text_w = "\n".join(text_content).rstrip('\n')
    
    pyperclip.copy(text_w) 
    var_show.set("")
    var_show.set("--%s数据复制成功!--" %name)
    print("--%s数据复制成功!--" %name)

#写入模块
#检索目标列最大行
def refresh_row(column):
    global old_wb
    old_wb = load_workbook('预应力钢筋.xlsx')
    a = write.rect_row(old_wb,column)
    var_row.set("")
    var_row.set(a)

def write_in_(old_wb, column1, first_row ,name1,name2):
    list1 = get_resluts(text3,result_num_write)
    list2 = get_resluts(text4,result_num_angle_write)
    global compare1
    global compare2
    if len(list1) == 0:
        print("横向数据为空！")
        var_show.set("")
        var_show.set("--横向数据为空!--")

    elif len(list2) == 0:
        print("纵向数据为空！")
        var_show.set("")
        var_show.set("--纵向向数据为空!--")

    else:
        if len(list1) != len(list2):
            print("--提示: 两侧数据个数不一致--\n--不过也可以写入--")
        #print("写入时比较列表的值为",compare1)
        #print("写入时list1 = ", list1)
        if compare1 != list1 or compare2 != list2:
            #把比较列表统一
            compare1.clear()
            for i in list1:
                compare1.append(i)
            compare2.clear()
            for i in list2:
                compare2.append(i)
            
            
            column2 = chr(ord(column1)+1)
            try:
                if int(list1[0]):
                    #print("横向表头为%s,已添加" %name1)
                    list1.insert(0,name1)

            except ValueError:
                print("此表已有名字")
            try:
                if int(list2[0]):
                    #print("横向表头为%s,已添加" %name2)
                    list2.insert(0,name2)
            except ValueError:
                print("此表已有名字")
            finally:       
                print("此时横向数据为：",list1)
                print("此时纵向数据为：",list2)
                first_row = int(first_row) + 2
                write.write_in(old_wb, column1, first_row, list1)
                write.write_in(old_wb, column2, first_row, list2)
                old_wb.save('预应力钢筋.xlsx')
                
                refresh_row(column2)
                refresh_row(column1)
                var_show.set("")
                var_show.set("--写入成功啦!--")
                global success
                success = True
                

        else:
            print("此表已写过")
            var_show.set("--此表已写入，请勿重复写入--")

#键盘事件
def key_function(event):
    #print("type = ", type(event))
    #print("event.char = ", event.char)
    #print("event.keycode = ", event.keycode)
    #print("event = ", event)
    pass


#tkinter初始化 
root = tk.Tk()
root.title("桥梁计算")
root.geometry("350x600")

#一直在前
root.attributes('-topmost', True)
#键盘事件
root.bind("<Key>", key_function)
#root.bind("<Control-a>",cut("left"))
#root.bind("<Control-d>",cut("right"))
_thread.start_new_thread(rect,(0.1, "a", "left"))
_thread.start_new_thread(rect,(0.1, "d", "right"))
_thread.start_new_thread(rect,(0.1, "x", "empty"))
#_thread.start_new_thread(rect,(0.1, "c", "copy"))
_thread.start_new_thread(rect,(0.1, "space", "count"))

change = True 
continue_rec = False
success = False

old_wb = load_workbook('预应力钢筋.xlsx')

var_radio = tk.StringVar()
var_radio.set('L')
var_check = tk.IntVar()
var_check.set(0)

r1 = tk.Radiobutton(root, text='左', variable=var_radio, value='L', command=from_left)
r2 = tk.Radiobutton(root, text='右', variable=var_radio, value='R', command=from_right)
c1 = tk.Checkbutton(root, text='连续识别',variable=var_check, onvalue=1, offvalue=0, command=continue_recogninze) 

#清空按钮
button0=tk.Button(root,text="一键清空",command=empty)

#识别模块
lablel_h = tk.Label(root, text="横向尺寸")
lablel_v = tk.Label(root, text="纵向尺寸")

text1 = tk.Text(root, width=20, height=9)
text2 = tk.Text(root, width=20, height=9)

button1=tk.Button(root,text="截图",command=lambda:cut("left"))
button2=tk.Button(root,text="识别",command=recognize)
button3=tk.Button(root,text="计算",command=calculate)
button3_1=tk.Button(root,text="除2",command=divide)
button4=tk.Button(root,text="截图",command=lambda:cut("right"))
button5=tk.Button(root,text="计算",command=calculate_angle)

#输出模块
label1 = tk.Label(root, text ="横向计算结果")
label2 = tk.Label(root, text ="纵向计算结果")

text3 = tk.Text(root, width=20, height=9)
text4 = tk.Text(root, width=20, height=9)

#复制按钮
button6=tk.Button(root,text="复制",command=lambda:copy_datas("横向",text3))
button7=tk.Button(root,text="复制",command=lambda:copy_datas("纵向",text4))

#写入模块
lablel3 = tk.Label(root, text ="表头")
lablel4 = tk.Label(root, text ="写入列：")
lablel5 = tk.Label(root, text ="写入行：")


var_h1 = tk.StringVar()
var_h2 = tk.StringVar()
var_h2.set("竖弯")
var_colunm = tk.StringVar()
var_colunm.set("A")
var_row = tk.StringVar()
var_row.set(write.rect_row(old_wb, "A"))
entry_h1 = tk.Entry(root, textvariable=var_h1, width=20)
entry_h2 = tk.Entry(root, textvariable=var_h2, width=20)
entry_column = tk.Entry(root, textvariable=var_colunm, width=5)
entry_row = tk.Entry(root, textvariable=var_row, width=5)
button_refresh = tk.Button(root,text="更新行数",command=lambda:refresh_row(var_colunm.get()))
button_write = tk.Button(root,text="写入数据",command=lambda:write_in_(old_wb, var_colunm.get(), var_row.get(),var_h1.get(),var_h2.get()))

var_show = tk.StringVar()
lablel_show = tk.Label(root, textvariable =var_show, fg="red")

#控件布局
r1.place(x=20, y=10)
r2.place(x=60, y=10)
c1.place(x=100,y=10)
button0.place(x=200,y=8)
lablel_h.place(x=20, y=40)
lablel_v.place(x=180, y=40)
text1.place(x=20,y=60)
text2.place(x=180,y=60)
button1.place(x=20,y=185)
#button2.place(x=60,y=200) 
button3_1.place(x=90,y=185)
button3.place(x=130,y=185) 
button4.place(x=180,y=185) 
button5.place(x=290,y=185) 
label1.place(x=20,y=240)
label2.place(x=180,y=240)
text3.place(x=20,y=260)
text4.place(x=180,y=260)
button6.place(x=109,y=380)
button7.place(x=269,y=380)

lablel3.place(x=20, y=400)
entry_h1.place(x=20,y=420)
entry_h2.place(x=180,y=420)
lablel4.place(x=20, y=450)
entry_column.place(x=80, y=450)
lablel5.place(x=20, y=470)
entry_row.place(x=80, y=470)
button_refresh.place(x=20,y=500)
button_write.place(x=20,y=530)

lablel_show.place(x=160,y=470)


root.mainloop()